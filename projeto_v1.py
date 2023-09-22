import openpyxl

primeiro = 'BOTAFOGO'
segundo = 'PALMEIRAS'
terceiro = 'GREMIO'
quarto = 'BRAGANTINO'
dezessete = 'SANTOS'
dezoito = 'VASCO'
dezenove = 'AMERICA-MG'
vinte = 'CORITIBA'
lista: list = [primeiro, segundo, terceiro, quarto, dezessete, dezoito, dezenove, vinte]
lista2: list = []

# Carregue o arquivo Excel (.xlsx)
arquivo = openpyxl.load_workbook('brasileirao.xlsx')

# Obtenha a primeira planilha
planilha = arquivo.active

def Contador(valor):
    contador = 0
    if valor[1] == primeiro:
        contador += 12
    elif valor[1] in lista:
        contador += 2
    else:
        pass

    if valor[2] == segundo:
        contador += 10
    elif valor[2] in lista:
        contador += 2
    else:
        pass

    if valor[3] == terceiro:
        contador += 9
    elif valor[3] in lista:
        contador += 2
    else:
        pass

    if valor[4] == quarto:
        contador += 8
    elif valor[4] in lista:
        contador += 2
    else:
        pass

    if valor[5] == dezessete:
        contador += 7
    elif valor[5] in lista:
        contador += 2
    else:
        pass

    if valor[6] == dezoito:
        contador += 6
    elif valor[6] in lista:
        contador += 2
    else:
        pass

    if valor[7] == dezenove:
        contador += 5
    elif valor[7] in lista:
        contador += 2
    else:
        pass

    if valor[8] == vinte:
        contador += 12
    elif valor[8] in lista:
        contador += 2
    else:
        pass
    return contador

# Comece da linha que você deseja (por exemplo, linha 2)
linha_desejada = 2  # Altere para o número da linha que deseja começar

jogador = 1
# Enquanto houver dados na planilha, continue
while planilha.cell(row=linha_desejada, column=1).value is not None:
    # Obtenha os valores das colunas 1 a 9 da linha desejada
    valores = [planilha.cell(row=linha_desejada, column=coluna).value for coluna in range(1, 10)]
    print(f'jogador' + str(jogador), {Contador(valores)})
    linha_desejada += 1
    jogador += 1
    lista2.append(Contador(valores))

print(lista2)
print(max(lista2))

